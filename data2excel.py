# -*- coding: utf-8 -*-
"""
Created on Mon Jan 15 09:36:46 2024

@author: Wilson

This file is used to read and save file according to the format of given 
printing setting data.

"""
import csv
import openpyxl

class XLS:
    def __init__(self, loadname, loadpage):
        wb = openpyxl.load_workbook(loadname)
        ws = wb[loadpage]
        ws = wb.active
        self.__workbook = openpyxl.Workbook()
        self.__booksheet = self.__workbook.active
        self.ws = list(ws.values)
    
    def save(self, filename, strip_color):
        # save new if not exist
        title = list(self.ws[0])
        extend = 0
        if 'OT_mean' not in title:
            title.append('OT_mean')
            extend += 1
        if 'MPM_mean' not in title:
            title.append('MPM_mean')
            extend += 1
        self.__booksheet.append(title)
        
        # check all index
        prop_idxs = {prop: title.index(prop) for prop in strip_color[0].keys()}
        for w in self.ws[1:]:
            w = list(w)
            for i in range(extend):
                w.append(0)
            for strip in strip_color:
                keylist = list(strip.keys())
                for key in keylist[:-1]:
                    if (strip[key] != w[prop_idxs[key]]):
                        break
                else:
                    w[prop_idxs[keylist[-1]]] = strip[keylist[-1]]
            self.__booksheet.append(w)
        self.__workbook.save(filename)
        self.__workbook.close()
        
def csv_read(filename):
    data = []
    with open(filename) as f:
        reader = csv.reader(f)
        for row in reader:
            data.append(float(row[1]))
    return data

if __name__ == '__main__':
    pass